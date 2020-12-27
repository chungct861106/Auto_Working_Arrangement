import pandas as pd
from gurobipy import *
import numpy as np
import openpyxl


def find_index(CSRs, name):
    index = 0
    for csr in CSRs:
        if csr[0] == name:
            return index
        index += 1

weekday = {
    "Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5, "Sun": 6
}

# File name
dataFile_path = "OR108-2_case00_data.xlsx"

# read demandInfo
monthDay = 31
demandInfo = pd.read_excel(dataFile_path, sheet_name='demand', header=None, skiprows=3)
demandInfo.columns = ["Time"] + [i for i in range(1, monthDay+1)]
demandInfo.set_index("Time", inplace=True)

# read shift
shiftInfo = pd.read_excel(dataFile_path, sheet_name='shifts', header=None, skiprows=2)
shiftInfo.drop([0, 1], axis=1, inplace=True)
shiftInfo.columns = ["Shift"] + list(demandInfo.index)
shiftInfo.set_index("Shift", inplace=True)

# read shift request
CSRs = pd.read_excel(dataFile_path, sheet_name='CSRs')

# read shift request
shiftRequest = pd.read_excel(dataFile_path, sheet_name='shift requests')

# read leave request
leaveRequest = pd.read_excel(dataFile_path, sheet_name='leave requests')

# read senior limit
senior_limit = pd.read_excel(dataFile_path, sheet_name='senior limit')

# read senior limit
manager_limit = pd.read_excel(dataFile_path, sheet_name='manager limit')


# define parameter
periodNum = len(shiftInfo.columns)
dateNum = len(demandInfo.columns)
CSRs_num = len(CSRs.index)

# change data structure
CSRs = CSRs.values.tolist()
demandInfo = demandInfo.values.tolist()
demandInfo = np.array(demandInfo)
demandInfo = np.transpose(demandInfo)
shiftInfo = shiftInfo.values.tolist()
shiftInfo = np.array(shiftInfo)
senior_limit = senior_limit.values.tolist()
manager_limit = manager_limit.values.tolist()


year = list()
level = list()
for csr in CSRs:
    year.append(csr[3])
    if csr[3] == "Manager":
        level.append(2)
    elif csr[3] == "Assistant Manager":
        level.append(1)
    else:
        level.append(0)
print(year)
# forming IP program

m = Model("solution")
w = list()  # lack data
N = list()  # supply data
x = list()  # CSRs allocation data


# Step 1: set variables

# supply variable (31x14)
for j in range(dateNum):
    N.append(list())
    for k in range(14):
        N[j].append(m.addVar(lb=0, vtype=GRB.INTEGER))
N = np.array(N)


# CSRs allocation data (40x31x14)
for i in range(CSRs_num):
    x.append(list())
    for j in range(dateNum):
        x[i].append(list())
        for k in range(14):
            x[i][j].append(m.addVar(lb=0, ub=1, vtype=GRB.INTEGER))

# lack variable (31x24)
for i in range(dateNum):
    w.append(list())
    for j in range(periodNum):
        w[i].append(m.addVar(lb=0, vtype=GRB.INTEGER))
m.update()


# Step 2: Set objective function

# oblective for total lack amount
objFunc = gurobipy.LinExpr()
for i in range(dateNum):
    for j in range(periodNum):
        objFunc += w[i][j]
m.setObjective(objFunc, GRB.MINIMIZE)   # minimize lack amount

# Step 3: Add in CSRs constraint

# Each csr assigned to only one shift each day
for csr in range(CSRs_num):
    for date in range(dateNum):
        CSRs_limits = gurobipy.LinExpr()
        for shift in x[csr][date]:
            CSRs_limits += shift
        m.addConstr(lhs=CSRs_limits, sense=GRB.EQUAL, rhs=1)

# Calculate to total supply amount
for date in range(dateNum):
    for shift in range(14):
        summantion = gurobipy.LinExpr()
        for csr in range(CSRs_num):
            summantion += x[csr][date][shift]
        m.addConstr(lhs=summantion, sense=GRB.EQUAL, rhs=N[date][shift])

# w define positve definition
supply = demandInfo - np.dot(N, shiftInfo)
for date in range(dateNum):
    for period in range(periodNum):
        positive_definite = gurobipy.LinExpr()
        positive_definite = w[date][period] - supply[date][period]
        m.addConstr(lhs=positive_definite, sense=GRB.GREATER_EQUAL, rhs=0)

# Month day off Constraint
for csr in range(CSRs_num):
    day_off_month_limits = gurobipy.LinExpr()
    for date in range(dateNum):
        day_off_month_limits += x[csr][date][13]
    m.addConstr(lhs=day_off_month_limits, sense=GRB.EQUAL, rhs=8)


# ------- Add in Request shift -------

request_shift = shiftRequest.values.tolist()

for request in request_shift:
    csr = find_index(CSRs, request[0])
    date = int(request[3].split("/")[1]) - 1
    constra = gurobipy.LinExpr()
    constra = x[csr][date][int(request[4]) - 1]
    assign = gurobipy.LinExpr()
    m.addConstr(lhs=constra, sense=GRB.EQUAL, rhs=1)

# --------- Add in Leave request -------
leave_request_shift = leaveRequest.values.tolist()
for request in leave_request_shift:
    csr = find_index(CSRs, request[0])
    date = str(request[3])
    if date.count("-") > 0:
        date = date.split("-")
        start = int(date[0].split("/")[1])
        end = int(date[1].split("/")[1])
        for d in range(start - 1, end):
            constra = gurobipy.LinExpr()
            constra = x[csr][d][13]
            m.addConstr(lhs=constra, sense=GRB.EQUAL, rhs=1)
    else:
        start = int(date.split("/")[1])-1
        constra = gurobipy.LinExpr()
        constra = x[csr][d][13]
        m.addConstr(lhs=constra, sense=GRB.EQUAL, rhs=1)

# --------- Week on shifts-------
for csr in range(CSRs_num):
    for start in range(dateNum-6):
        # Day off limitsV
        dayoff_shift_limits = gurobipy.LinExpr()
        for date in range(start, start + 7):
            dayoff_shift_limits += x[csr][date][13]
        m.addConstr(lhs=dayoff_shift_limits, sense=GRB.GREATER_EQUAL, rhs=1)

        # Night limitsV
        night_shift_limits = gurobipy.LinExpr()
        for date in range(start, start + 7):
            for shift in range(10, 13):
                night_shift_limits += x[csr][date][shift]
        m.addConstr(lhs=night_shift_limits, sense=GRB.LESS_EQUAL, rhs=1)

        # Afternoon limitsV
        Afternoon_shift_limits = gurobipy.LinExpr()
        for date in range(start, start + 7):
            for shift in range(6, 10):
                Afternoon_shift_limits += x[csr][date][shift]
        m.addConstr(lhs=Afternoon_shift_limits, sense=GRB.LESS_EQUAL, rhs=2)

for limit in manager_limit:
    day = limit[0].date().day
    limit_shift = []
    if limit[1] == "night":
        limit_shift = [10, 11, 12]
    elif limit[1] == "afternoon":
        limit_shift = [6, 7, 8, 9]
    else:
        limit_shift = [0, 1, 2, 3, 4, 5]
    limit_level = 0
    if limit[2] == "Manager":
        limit_level = 2
    else:
        limit_level = 1
    manager = gurobipy.LinExpr()
    for shift in limit_shift:
        for csr in range(CSRs_num):
            if level[csr] >= limit_level:
                manager += x[csr][date][shift]
    m.addConstr(lhs=manager, sense=GRB.GREATER_EQUAL, rhs=int(limit[3]))

first = weekday["Mon"]
for limit in senior_limit:
    limit_weekday = weekday[limit[0]] + first
    times = limit[1].split("-")
    start_time = times[0].split(":")
    end_time = times[1].split(":")
    start_index = 2 * (int(start_time[0]) - 9)
    if start_time[1] == "30":
        start_index += 1
    end_index = 2 * (int(end_time[0]) - 9) - 1
    if end_time[1] == "30":
        end_index += 1
    limit_shifts = list()
    for period in range(start_index, end_index + 1):
        this_limit_shift = list()
        for shift in range(14):
            if shiftInfo[shift][period] == 1:
                this_limit_shift.append(shift)
        limit_shifts.append(this_limit_shift)
    last_sum = 0
    for shifts in limit_shifts:
        if sum(shifts) == last_sum:
            continue
        else:
            last_sum = sum(shifts)
        n = 0
        while limit_weekday + n * 7 < dateNum:
            senior = gurobipy.LinExpr()
            total_csr = gurobipy.LinExpr()
            for csr in range(CSRs_num):
                for shift in shifts:
                    if year[csr] >= limit[3]:
                        senior += x[csr][limit_weekday + n * 7][shift]
                    total_csr += x[csr][limit_weekday + n * 7][shift]
            n += 1
            m.addConstr(lhs=senior, sense=GRB.GREATER_EQUAL, rhs=float(limit[2]) * total_csr)
m.optimize()
workbook = openpyxl.load_workbook(dataFile_path)
worksheet = workbook.get_sheet_by_name('CSRs test place')
start_row = 2
for csr in range(40):
    for date in range(dateNum):
        for shift in range(14):
            if x[csr][date][shift].x == 1:
                worksheet.cell(row=start_row + csr, column=3 + date).value = (shift + 1) % 14
                break

try:
    worksheet = workbook.get_sheet_by_name('Output Data')
    workbook.remove_sheet(worksheet)
except:
    pass
workbook.create_sheet('Output Data')
worksheet = workbook.get_sheet_by_name('Output Data')
headline = ["Date\shift"]
for num in range(1, 14, 1):
    headline.append(num)
headline.append(0)
worksheet.append(headline)
now_date = 1
Out_data = []
for rows in N:
    data = []
    for value in rows:
        data.append(value.x)
    Out_data.append(data)
    worksheet.append([now_date]+data)
    now_date += 1

Out_data = np.array(Out_data)
Supply = demandInfo-np.dot(Out_data, shiftInfo)

try:
    worksheet = workbook.get_sheet_by_name('Output Demand')
    workbook.remove_sheet(worksheet)
except:
    pass
workbook.create_sheet('Output Demand')
worksheet = workbook.get_sheet_by_name('Output Demand')
headline = ["Date\period"]
for num in range(1, 25, 1):
    headline.append(num)
now_date = 1
worksheet.append(headline)
for rows in Supply:
    data = []
    for num in rows:
        data.append(num)
    worksheet.append([now_date]+data)
    now_date += 1
workbook.save(dataFile_path)
